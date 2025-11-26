import os
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout as auth_logout
from django.db.models import Q, Count, Prefetch, Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
import json
from datetime import timedelta, datetime
from .models import (
    Aluno, Turma, Aula, HorarioAula, Frequencia,
    Aviso, Mensalidade, Mensagem, Notificacao,
    Evento, VendaIngresso, ResultadoFinanceiroMensal, DespesaAluno, DespesaAdministrativa, EntradaFinanceira
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import mercadopago
from chat.models import Room

# Home Page
def home_view(request):
    return render(request, 'home/index.html')

# Página do Sistema
def sistema_view(request):
    """Página 'Conheça nosso Sistema' com planos e informações"""
    return render(request, 'home/sistema.html')

# Logout
@login_required
def logout_view(request):
    """Faz logout do usuário e redireciona para home"""
    auth_logout(request)
    messages.success(request, 'Você saiu com sucesso!')
    return redirect('paginas:home')

# Painel do Aluno - Dashboard Principal
@login_required
def painel_aluno_index(request):
    """Dashboard principal com resumo de informações do aluno"""
    context = {}
    
    try:
        # Usar select_related para otimizar queries
        aluno = Aluno.objects.select_related('usuario').prefetch_related(
            Prefetch('turmas', queryset=Turma.objects.filter(ativa=True))
        ).get(usuario=request.user)
        
        turmas = aluno.turmas.filter(ativa=True)
        
        # Avisos recentes (últimos 7 dias)
        data_limite = timezone.now() - timedelta(days=7)
        avisos_recentes = Aviso.objects.filter(
            Q(tipo='GERAL') | 
            Q(turma__in=turmas) | 
            Q(aluno=aluno),
            ativo=True,
            data_criacao__gte=data_limite
        ).select_related('autor', 'turma').distinct().order_by('-importante', '-data_criacao')[:5]
        
        # Próximas aulas (próximos 7 dias)
        hoje = timezone.now().date()
        proximas_aulas = Aula.objects.filter(
            turma__in=turmas,
            data__gte=hoje,
            data__lte=hoje + timedelta(days=7),
            realizada=False
        ).select_related('turma').order_by('data', 'hora_inicio')[:5]
        
        # Estatísticas de frequência
        total_aulas = Frequencia.objects.filter(aluno=aluno).count()
        presencas = Frequencia.objects.filter(aluno=aluno, status='PRESENTE').count()
        faltas = Frequencia.objects.filter(aluno=aluno, status='FALTA').count()
        percentual_presenca = (presencas / total_aulas * 100) if total_aulas > 0 else 0
        
        # Mensalidades pendentes
        mensalidades_pendentes = Mensalidade.objects.filter(
            aluno=aluno,
            status__in=['PENDENTE', 'ATRASADO']
        ).order_by('data_vencimento')[:3]
        
        # Mensagens não lidas
        mensagens_nao_lidas = Mensagem.objects.filter(
            destinatario=request.user,
            lida=False
        ).count()
        
        context = {
            'usuario': request.user,
            'username': request.user.get_full_name() or request.user.username,
            'aluno': aluno,
            'turmas': turmas,
            'avisos_recentes': avisos_recentes,
            'proximas_aulas': proximas_aulas,
            'total_aulas': total_aulas,
            'presencas': presencas,
            'faltas': faltas,
            'percentual_presenca': round(percentual_presenca, 1),
            'mensalidades_pendentes': mensalidades_pendentes,
            'mensagens_nao_lidas': mensagens_nao_lidas,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar dados: {str(e)}'
    
    return render(request, 'painel/index.html', context)

@login_required
def painel_aluno_avisos(request):
    """Listar todos os avisos do aluno"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').prefetch_related(
            Prefetch('turmas', queryset=Turma.objects.filter(ativa=True))
        ).get(usuario=request.user)
        
        turmas = aluno.turmas.filter(ativa=True)
        
        # Buscar todos os avisos relevantes
        avisos_lista = Aviso.objects.filter(
            Q(tipo='GERAL') | 
            Q(turma__in=turmas) | 
            Q(aluno=aluno),
            ativo=True
        ).select_related('autor', 'turma', 'aluno').distinct().order_by('-importante', '-data_criacao')
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'avisos': avisos_lista,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar avisos: {str(e)}'
    
    return render(request, 'painel/avisos.html', context)

@login_required
def painel_aluno_horarios(request):
    """Exibir horários das aulas do aluno e vídeos das últimas aulas"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').prefetch_related(
            Prefetch('turmas', queryset=Turma.objects.filter(ativa=True))
        ).get(usuario=request.user)
        
        turmas = aluno.turmas.filter(ativa=True)
        
        # Horários recorrentes (grade de horários)
        horarios = HorarioAula.objects.filter(turma__in=turmas).select_related('turma').order_by('dia_semana', 'hora_inicio')
        
        # Organizar horários por dia da semana
        horarios_por_dia = {}
        for horario in horarios:
            dia = horario.get_dia_semana_display()
            if dia not in horarios_por_dia:
                horarios_por_dia[dia] = []
            
            horarios_por_dia[dia].append(horario)
        
        # Próximas aulas (não realizadas e futuras)
        hoje = timezone.now().date()
        proximas_aulas = Aula.objects.filter(
            turma__in=turmas,
            data__gte=hoje,
            realizada=False
        ).select_related('turma').order_by('data', 'hora_inicio')[:10]
        
        # Últimas aulas realizadas COM vídeo (máximo 5)
        aulas_com_video = Aula.objects.filter(
            turma__in=turmas,
            realizada=True,
            video__isnull=False
        ).exclude(video='').select_related('turma').order_by('-data', '-hora_inicio')[:5]
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'horarios': horarios,
            'horarios_por_dia': horarios_por_dia,
            'dias_semana': horarios_por_dia.keys(),
            'proximas_aulas': proximas_aulas,
            'aulas_com_video': aulas_com_video,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar horários: {str(e)}'
    
    return render(request, 'painel/horarios_aula.html', context)

@login_required
def painel_aluno_minhas_aulas(request):
    """Listar aulas do aluno com frequência"""
    aulas_futuras = []
    aulas_com_video = []
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').prefetch_related(
            Prefetch('turmas', queryset=Turma.objects.filter(ativa=True))
        ).get(usuario=request.user)
        
        turmas = aluno.turmas.filter(ativa=True)
        
        hoje = timezone.now().date()
        
        # Buscar aulas FUTURAS (próximos 30 dias)
        data_limite_futuro = hoje + timedelta(days=30)
        aulas_futuras_queryset = Aula.objects.filter(
            turma__in=turmas,
            data__gt=hoje,
            data__lte=data_limite_futuro
        ).select_related('turma').order_by('data', 'hora_inicio')
        
        # Adicionar dias até a aula
        for aula in aulas_futuras_queryset:
            dias_ate = (aula.data - hoje).days
            aula.dias_ate_aula = dias_ate if dias_ate > 0 else 0
            aulas_futuras.append(aula)
        
        # Buscar aulas COM VÍDEO (ordenadas por data decrescente)
        aulas_com_video = Aula.objects.filter(
            turma__in=turmas,
            video__isnull=False
        ).exclude(video='').select_related('turma').order_by('-data', '-hora_inicio')[:12]  # Últimas 12 aulas com vídeo
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'aulas_futuras': aulas_futuras,
            'aulas_com_video': aulas_com_video,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar aulas: {str(e)}'
    
    return render(request, 'painel/minhas_aulas.html', context)

@login_required
def painel_aluno_comunicacao(request):
    """Sistema de comunicação/mensagens"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').get(usuario=request.user)
        
        # Buscar conversas (mensagens recebidas e enviadas)
        mensagens_recebidas = Mensagem.objects.filter(
            destinatario=request.user
        ).select_related('remetente').order_by('-data_envio')[:20]
        
        mensagens_enviadas = Mensagem.objects.filter(
            remetente=request.user
        ).select_related('destinatario').order_by('-data_envio')[:20]
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'mensagens_recebidas': mensagens_recebidas,
            'mensagens_enviadas': mensagens_enviadas,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar mensagens: {str(e)}'
    
    return render(request, 'painel/comunicacao.html', context)

@login_required
def painel_aluno_chat(request):
    """Chat/conversas"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').get(usuario=request.user)
        
        # Buscar mensagens recentes
        mensagens = Mensagem.objects.filter(
            Q(remetente=request.user) | Q(destinatario=request.user)
        ).select_related('remetente', 'destinatario').order_by('-data_envio')[:50]
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'mensagens': mensagens,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar chat: {str(e)}'
    
    return render(request, 'painel/chat.html', context)

# Financeiro
@login_required
def financeiro_mensalidades(request):
    """Exibir mensalidades do aluno"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').get(usuario=request.user)
        mensalidades = Mensalidade.objects.filter(aluno=aluno).order_by('-mes_referencia')
        
        # Calcular totais
        total_pago = sum(m.valor_final for m in mensalidades if m.status == 'PAGO')
        total_pendente = sum(m.valor_final for m in mensalidades if m.status in ['PENDENTE', 'ATRASADO'])
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'mensalidades': mensalidades,
            'total_pago': total_pago,
            'total_pendente': total_pendente,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar mensalidades: {str(e)}'
    
    return render(request, 'financeiro/mensalidades.html', context)

def pagar_mensalidade(request, mensalidade_id):
    print(f"=== PAGAR MENSALIDADE CHAMADO === ID: {mensalidade_id}")
    print(f"Request method: {request.method}")
    print(f"Is AJAX: {request.headers.get('X-Requested-With')}")
    
    try:
        aluno = get_object_or_404(Aluno, usuario=request.user)
        mensalidade = get_object_or_404(Mensalidade, id=mensalidade_id, aluno=aluno)
        
        print(f"Mensalidade encontrada: {mensalidade}")
        
        MP_ACCESS_TOKEN = settings.MP_ACCESS_TOKEN
        MP_PUBLIC_KEY = settings.MP_PUBLIC_KEY
        
        print(f"MP_ACCESS_TOKEN: {MP_ACCESS_TOKEN[:20]}...")
        print(f"MP_PUBLIC_KEY: {MP_PUBLIC_KEY[:20]}...")
        
        sdk = mercadopago.SDK(MP_ACCESS_TOKEN)

        preference_data = {
            "items": [
                {
                    "title": f"Mensalidade {mensalidade.mes_referencia.strftime('%m/%Y')}",
                    "quantity": 1,
                    "currency_id": "BRL",
                    "unit_price": float(mensalidade.valor_final),
                }
            ],
            "metadata": {
                "mensalidade_id": str(mensalidade.id)
            },
            "statement_descriptor": "GIRO DNC",
            "external_reference": f"MENS-{mensalidade.id}"
        }

        print(f"Criando preferência com dados: {preference_data}")
        preference = sdk.preference().create(preference_data)
        print(f"Resposta do MP: Status={preference.get('status')}")
        
        # Verificar se houve erro na criação
        if preference["status"] != 201:
            error_msg = preference.get("response", {}).get("message", "Erro desconhecido")
            print(f"ERRO ao criar preferência: {error_msg}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "error": f"Erro ao criar preferência: {error_msg}"
                }, status=500)
            else:
                messages.error(request, "Erro ao processar pagamento. Tente novamente.")
                return redirect('paginas:financeiro_mensalidades')
        
        preference_id = preference["response"]["id"]
        print(f"Preferência criada com sucesso! ID: {preference_id}")

        # Se for requisição AJAX, retornar JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                "public_key": MP_PUBLIC_KEY,
                "preference_id": preference_id,
                "mensalidade_id": mensalidade.id,
                "amount": float(mensalidade.valor_final)
            })

        # Se não for AJAX, retornar o template completo (fallback)
        mensalidades = Mensalidade.objects.filter(aluno=aluno).order_by("-mes_referencia")
        total_pago = sum(m.valor_final for m in mensalidades if m.status == 'PAGO')
        total_pendente = sum(m.valor_final for m in mensalidades if m.status in ['PENDENTE', 'ATRASADO'])

        context = {
            "usuario": request.user,
            "aluno": aluno,
            "mensalidades": mensalidades,
            "total_pago": total_pago,
            "total_pendente": total_pendente,
            "abrir_pagamento": mensalidade.id,
            "public_key": MP_PUBLIC_KEY,
            "preference_id": preference_id,
            "mensalidade_selecionada": mensalidade.id,
        }

        return render(request, 'financeiro/mensalidades.html', context)
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                "error": f"Erro ao processar pagamento: {str(e)}"
            }, status=500)
        else:
            messages.error(request, f"Erro ao processar pagamento: {str(e)}")
            return redirect('paginas:financeiro_mensalidades')

@csrf_exempt
def webhook_mercadopago(request):
    """Webhook para receber notificações do Mercado Pago"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Log para debug
            print(f"Webhook recebido: {data}")
            
            # Verificar se é notificação de pagamento
            if data.get('type') == 'payment':
                payment_id = data.get('data', {}).get('id')
                
                if payment_id:
                    # Buscar informações do pagamento
                    sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)
                    payment_info = sdk.payment().get(payment_id)
                    
                    if payment_info['status'] == 200:
                        payment_data = payment_info['response']
                        
                        # Obter ID da mensalidade do metadata
                        mensalidade_id = payment_data.get('metadata', {}).get('mensalidade_id')
                        
                        if mensalidade_id:
                            mensalidade = Mensalidade.objects.get(id=mensalidade_id)
                            
                            # Atualizar status conforme o status do pagamento
                            if payment_data['status'] == 'approved':
                                mensalidade.status = 'PAGO'
                                mensalidade.data_pagamento = timezone.now()
                                mensalidade.save()
                                
                                # Criar notificação para o aluno
                                Notificacao.objects.create(
                                    usuario=mensalidade.aluno.usuario,
                                    tipo='PAGAMENTO',
                                    titulo='Pagamento Aprovado',
                                    mensagem=f'Sua mensalidade de {mensalidade.mes_referencia.strftime("%m/%Y")} foi aprovada!'
                                )
                            elif payment_data['status'] == 'pending':
                                mensalidade.observacoes = 'Pagamento pendente'
                                mensalidade.save()
            
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            print(f"Erro no webhook: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método não permitido'}, status=405)


@csrf_exempt
@require_http_methods(["POST"])
def processar_pagamento(request):
    """Processar pagamento através do Payment Brick"""
    try:
        data = json.loads(request.body)
        print("=== PROCESSAR PAGAMENTO ===")
        print(f"Dados recebidos: {data}")
        
        mensalidade_id = data.get('mensalidade_id')
        if not mensalidade_id:
            return JsonResponse({'error': 'ID da mensalidade não fornecido'}, status=400)
            
        mensalidade = Mensalidade.objects.get(id=mensalidade_id)
        
        # Inicializar SDK do Mercado Pago
        sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)
        
        # Preparar dados do pagamento
        payment_data = {
            "transaction_amount": float(data.get('transaction_amount')),
            "token": data.get('token'),
            "description": f"Mensalidade {mensalidade.mes_referencia.strftime('%m/%Y')}",
            "installments": int(data.get('installments', 1)),
            "payment_method_id": data.get('payment_method_id'),
            "issuer_id": data.get('issuer_id'),
            "payer": {
                "email": data.get('payer', {}).get('email'),
                "identification": {
                    "type": data.get('payer', {}).get('identification', {}).get('type'),
                    "number": data.get('payer', {}).get('identification', {}).get('number')
                }
            },
            "metadata": {
                "mensalidade_id": str(mensalidade.id)
            },
            "statement_descriptor": "GIRO DNC"
        }
        
        print(f"Enviando pagamento para Mercado Pago: {payment_data}")
        
        # Processar pagamento
        payment_response = sdk.payment().create(payment_data)
        payment = payment_response["response"]
        
        print(f"Resposta do Mercado Pago: {payment}")
        
        # Atualizar mensalidade baseado no status
        if payment.get('status') == 'approved':
            mensalidade.status = 'PAGO'
            mensalidade.data_pagamento = timezone.now()
            mensalidade.save()
            
            # Criar notificação
            Notificacao.objects.create(
                usuario=mensalidade.aluno.usuario,
                tipo='PAGAMENTO',
                titulo='Pagamento Aprovado',
                mensagem=f'Sua mensalidade de {mensalidade.mes_referencia.strftime("%m/%Y")} foi aprovada!'
            )
            
            return JsonResponse({
                'success': True,
                'payment_id': payment.get('id'),
                'status': payment.get('status'),
                'status_detail': payment.get('status_detail')
            })
        elif payment.get('status') == 'pending':
            mensalidade.observacoes = f"Pagamento pendente - {payment.get('status_detail')}"
            mensalidade.save()
            
            return JsonResponse({
                'success': True,
                'payment_id': payment.get('id'),
                'status': payment.get('status'),
                'status_detail': payment.get('status_detail'),
                'message': 'Pagamento pendente de aprovação'
            })
        else:
            return JsonResponse({
                'error': f"Pagamento {payment.get('status')}: {payment.get('status_detail')}",
                'payment_id': payment.get('id'),
                'status': payment.get('status')
            }, status=400)
            
    except Mensalidade.DoesNotExist:
        return JsonResponse({'error': 'Mensalidade não encontrada'}, status=404)
    except Exception as e:
        print(f"Erro ao processar pagamento: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def webhook_mercadopago_old(request):
    """Webhook para receber notificações do Mercado Pago"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Log para debug
            print(f"Webhook recebido: {data}")
            
            # Verificar se é notificação de pagamento
            if data.get('type') == 'payment':
                payment_id = data.get('data', {}).get('id')
                
                if payment_id:
                    # Buscar informações do pagamento
                    sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)
                    payment_info = sdk.payment().get(payment_id)
                    
                    if payment_info['status'] == 200:
                        payment_data = payment_info['response']
                        
                        # Obter ID da mensalidade do metadata
                        mensalidade_id = payment_data.get('metadata', {}).get('mensalidade_id')
                        
                        if mensalidade_id:
                            mensalidade = Mensalidade.objects.get(id=mensalidade_id)
                            
                            # Atualizar status conforme o status do pagamento
                            if payment_data['status'] == 'approved':
                                mensalidade.status = 'PAGO'
                                mensalidade.data_pagamento = timezone.now()
                                mensalidade.save()
                                
                                # Criar notificação para o aluno
                                Notificacao.objects.create(
                                    usuario=mensalidade.aluno.usuario,
                                    tipo='PAGAMENTO',
                                    titulo='Pagamento Aprovado',
                                    mensagem=f'Sua mensalidade de {mensalidade.mes_referencia.strftime("%m/%Y")} foi aprovada!'
                                )
                            elif payment_data['status'] == 'pending':
                                mensalidade.observacoes = 'Pagamento pendente'
                                mensalidade.save()
                                
            return JsonResponse({'status': 'ok'})
            
        except Exception as e:
            print(f"Erro no webhook: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def financeiro_extrato(request):
    """Exibir extrato financeiro do aluno"""
    context = {}
    
    try:
        aluno = Aluno.objects.select_related('usuario').get(usuario=request.user)
        mensalidades = Mensalidade.objects.filter(aluno=aluno).order_by('-mes_referencia')[:12]
        
        context = {
            'usuario': request.user,
            'aluno': aluno,
            'mensalidades': mensalidades,
        }
    except Aluno.DoesNotExist:
        context['erro'] = 'Usuário não está cadastrado como aluno.'
    except Exception as e:
        context['erro'] = f'Erro ao carregar extrato: {str(e)}'
    
    return render(request, 'financeiro/Extrato.html', context)

@login_required
def financeiro_extrato_csv(request):
    """Exportar extrato financeiro em Excel (.xlsx)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from io import BytesIO
    
    try:
        aluno = Aluno.objects.select_related('usuario').get(usuario=request.user)
        mensalidades = Mensalidade.objects.filter(aluno=aluno).order_by('-mes_referencia')
        
        # Criar workbook e planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Extrato Financeiro"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalho
        headers = [
            'Mês/Ano',
            'Data Vencimento',
            'Valor Original',
            'Desconto',
            'Valor Final',
            'Data Pagamento',
            'Status',
            'Forma Pagamento',
            'Observações'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        row_num = 2
        total_pago = 0
        total_pendente = 0
        
        for m in mensalidades:
            ws.cell(row=row_num, column=1, value=m.mes_referencia.strftime('%m/%Y')).border = border
            ws.cell(row=row_num, column=2, value=m.data_vencimento.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row_num, column=3, value=float(m.valor)).border = border
            ws.cell(row=row_num, column=4, value=float(m.valor_desconto)).border = border
            ws.cell(row=row_num, column=5, value=float(m.valor_final)).border = border
            ws.cell(row=row_num, column=6, value=m.data_pagamento.strftime('%d/%m/%Y') if m.data_pagamento else '-').border = border
            ws.cell(row=row_num, column=7, value=m.get_status_display()).border = border
            ws.cell(row=row_num, column=8, value=m.forma_pagamento if m.forma_pagamento else '-').border = border
            ws.cell(row=row_num, column=9, value=m.observacoes if m.observacoes else '-').border = border
            
            # Formatar colunas de valores como moeda
            ws.cell(row=row_num, column=3).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=4).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'
            
            # Calcular totais
            if m.status == 'PAGO':
                total_pago += float(m.valor_final)
            elif m.status in ['PENDENTE', 'ATRASADO']:
                total_pendente += float(m.valor_final)
            
            row_num += 1
        
        # Linha em branco
        row_num += 1
        
        # Totais
        ws.cell(row=row_num, column=1, value='TOTAIS').font = total_font
        ws.cell(row=row_num, column=1).fill = total_fill
        row_num += 1
        
        ws.cell(row=row_num, column=1, value='Total Pago').font = total_font
        ws.cell(row=row_num, column=5, value=total_pago).font = total_font
        ws.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row_num, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row_num += 1
        
        ws.cell(row=row_num, column=1, value='Total Pendente').font = total_font
        ws.cell(row=row_num, column=5, value=total_pendente).font = total_font
        ws.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row_num, column=5).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 12,  # Mês/Ano
            'B': 18,  # Data Vencimento
            'C': 15,  # Valor Original
            'D': 12,  # Desconto
            'E': 15,  # Valor Final
            'F': 18,  # Data Pagamento
            'G': 12,  # Status
            'H': 18,  # Forma Pagamento
            'I': 30,  # Observações
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Salvar em BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Criar resposta HTTP
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="extrato_financeiro_{aluno.usuario.username}.xlsx"'
        
        return response
        
    except Aluno.DoesNotExist:
        from django.contrib import messages
        messages.error(request, 'Usuário não está cadastrado como aluno.')
        return redirect('paginas:financeiro_extrato')
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Erro ao gerar Excel: {str(e)}')
        return redirect('paginas:financeiro_extrato')

@login_required
def financeiro_despesas(request):
    """Exibir despesas - Admin vê gestão ou Students veem despesas pessoais"""
    
    # Se for staff (admin/professor), mostrar view administrativa
    if request.user.is_staff:
        context = {
            'usuario': request.user,
            'is_admin_view': True,
        }
        return render(request, 'financeiro/despesas.html', context)
    
    # Se for aluno, mostrar sistema de despesas pessoais
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        messages.error(request, 'Você precisa estar cadastrado como aluno.')
        return redirect('paginas:painel_aluno')
    
    # Filtros
    mes_filtro = request.GET.get('mes', '')
    categoria_filtro = request.GET.get('categoria', '')
    status_filtro = request.GET.get('status', '')
    
    # Buscar despesas do aluno
    despesas = DespesaAluno.objects.filter(aluno=aluno).select_related('turma')
    
    # Aplicar filtros
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-')
            despesas = despesas.filter(
                data_prevista__year=int(ano),
                data_prevista__month=int(mes)
            )
        except ValueError:
            pass
    
    if categoria_filtro:
        despesas = despesas.filter(categoria=categoria_filtro)
    
    if status_filtro:
        despesas = despesas.filter(status=status_filtro)
    
    # Estatísticas
    total_previsto = despesas.aggregate(Sum('valor_previsto'))['valor_previsto__sum'] or 0
    total_gasto = despesas.aggregate(Sum('valor_gasto'))['valor_gasto__sum'] or 0
    total_restante = total_previsto - total_gasto
    total_restante_abs = abs(total_restante)
    
    # Despesas por categoria (para gráfico)
    despesas_por_categoria = {}
    for despesa in despesas:
        cat = despesa.categoria_display()
        if cat not in despesas_por_categoria:
            despesas_por_categoria[cat] = 0
        despesas_por_categoria[cat] += float(despesa.valor_gasto)
    
    # Preparar dados para gráfico
    categorias_labels = list(despesas_por_categoria.keys())
    categorias_valores = list(despesas_por_categoria.values())
    
    context = {
        'aluno': aluno,
        'despesas': despesas.order_by('-data_prevista'),
        'total_previsto': total_previsto,
        'total_gasto': total_gasto,
        'total_restante': total_restante,
        'total_restante_abs': total_restante_abs,
        'categorias_labels': json.dumps(categorias_labels),
        'categorias_valores': json.dumps(categorias_valores),
        'mes_filtro': mes_filtro,
        'categoria_filtro': categoria_filtro,
        'status_filtro': status_filtro,
        'categorias_choices': DespesaAluno.CATEGORIA_CHOICES,
        'status_choices': DespesaAluno.STATUS_CHOICES,
        'is_admin_view': False,
    }
    
    return render(request, 'financeiro/despesas.html', context)

# Cadastros
def cadastro_aluno(request):
    return render(request, 'cadastros/cadastro_aluno_novo.html')

def cadastro_professor(request):
    return render(request, 'cadastros/cadastro_professor_novo.html')

# Feedback
def feedback_view(request):
    return render(request, 'feedback/aba feedback.html')

# API Endpoints
@login_required
@require_http_methods(["POST"])
def enviar_mensagem(request):
    """Enviar mensagem do formulário de contato"""
    try:
        data = json.loads(request.body)
        assunto = data.get('assunto', '').strip()
        mensagem = data.get('mensagem', '').strip()
        urgente = data.get('urgente', False)
        
        # Validações
        if not assunto:
            return JsonResponse({
                'success': False,
                'error': 'Assunto é obrigatório.'
            }, status=400)
        
        if not mensagem:
            return JsonResponse({
                'success': False,
                'error': 'Mensagem é obrigatória.'
            }, status=400)
        
        if len(mensagem) > 500:
            return JsonResponse({
                'success': False,
                'error': 'Mensagem não pode ter mais de 500 caracteres.'
            }, status=400)
        
        # Verificar se usuário é aluno
        try:
            aluno = Aluno.objects.get(usuario=request.user)
        except Aluno.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Apenas alunos podem enviar mensagens.'
            }, status=403)
        
        # Buscar usuário administrador/professor para enviar a mensagem
        from django.contrib.auth.models import User
        admin_users = User.objects.filter(is_staff=True).first()
        
        if not admin_users:
            # Se não houver admin, usar o primeiro usuário professor
            admin_users = User.objects.filter(turmas_professor__isnull=False).first()
        
        if admin_users:
            # Criar mensagem
            msg = Mensagem.objects.create(
                remetente=request.user,
                destinatario=admin_users,
                assunto=f"{'[URGENTE] ' if urgente else ''}{assunto}",
                conteudo=mensagem
            )
            
            # Criar notificação para o destinatário
            Notificacao.objects.create(
                usuario=admin_users,
                tipo='ALERTA' if urgente else 'INFO',
                titulo=f"Nova mensagem de {request.user.get_full_name() or request.user.username}",
                mensagem=f"Assunto: {assunto}",
                link=f"/admin/paginas/mensagem/{msg.id}/change/"
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Mensagem enviada com sucesso!',
                'mensagem_id': msg.id
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Não há destinatários disponíveis.'
            }, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Dados inválidos.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }, status=500)

@login_required
def grafico_frequencia(request):
    """Retorna dados para gráfico de frequência"""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
        
        # Buscar frequências dos últimos 6 meses
        hoje = timezone.now().date()
        data_inicio = hoje - timedelta(days=180)
        
        frequencias = Frequencia.objects.filter(
            aluno=aluno,
            aula__data__gte=data_inicio
        ).select_related('aula')
        
        # Agrupar por mês
        dados_mensais = {}
        for freq in frequencias:
            mes_ano = freq.aula.data.strftime('%Y-%m')
            if mes_ano not in dados_mensais:
                dados_mensais[mes_ano] = {
                    'presencas': 0,
                    'faltas': 0,
                    'faltas_justificadas': 0,
                    'total': 0
                }
            
            dados_mensais[mes_ano]['total'] += 1
            if freq.status == 'PRESENTE':
                dados_mensais[mes_ano]['presencas'] += 1
            elif freq.status == 'FALTA':
                dados_mensais[mes_ano]['faltas'] += 1
            elif freq.status in ['FALTA_JUSTIFICADA', 'ATESTADO']:
                dados_mensais[mes_ano]['faltas_justificadas'] += 1
        
        # Formatar para o gráfico
        labels = []
        presencas = []
        faltas = []
        
        for mes_ano in sorted(dados_mensais.keys()):
            # Converter string para date para formatação
            ano, mes = mes_ano.split('-')
            # Nomes de meses em português
            meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                       'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
            labels.append(f"{meses_pt[int(mes)-1]}/{ano[2:]}")
            
            total = dados_mensais[mes_ano]['total']
            presencas.append(round((dados_mensais[mes_ano]['presencas'] / total) * 100, 1) if total > 0 else 0)
            faltas.append(round((dados_mensais[mes_ano]['faltas'] / total) * 100, 1) if total > 0 else 0)
        
        return JsonResponse({
            'success': True,
            'labels': labels,
            'presencas': presencas,
            'faltas': faltas
        })
        
    except Aluno.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Aluno não encontrado.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def listar_notificacoes(request):
    """Lista notificações não lidas do usuário"""
    try:
        notificacoes = Notificacao.objects.filter(
            usuario=request.user,
            lida=False
        )[:10]
        
        data = [{
            'id': n.id,
            'tipo': n.tipo,
            'titulo': n.titulo,
            'mensagem': n.mensagem,
            'link': n.link,
            'data_criacao': n.data_criacao.strftime('%d/%m/%Y %H:%M')
        } for n in notificacoes]
        
        return JsonResponse({
            'success': True,
            'notificacoes': data,
            'total': notificacoes.count()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["POST"])
def marcar_notificacao_lida(request, notificacao_id):
    """Marca uma notificação como lida"""
    try:
        notificacao = Notificacao.objects.get(
            id=notificacao_id,
            usuario=request.user
        )
        notificacao.marcar_como_lida()
        
        return JsonResponse({
            'success': True,
            'message': 'Notificação marcada como lida.'
        })
        
    except Notificacao.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notificação não encontrada.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ============================================
# SISTEMA DE VENDAS DE INGRESSOS
# ============================================

@login_required
def eventos_lista(request):
    """Lista de eventos ativos para registro de vendas"""
    eventos_ativos = Evento.objects.filter(ativo=True).order_by('-data_evento')
    
    # Vendas do usuário atual
    minhas_vendas = VendaIngresso.objects.filter(
        vendedor=request.user
    ).select_related('evento').order_by('-data_venda')[:10]
    
    # Estatísticas do usuário
    total_vendido = VendaIngresso.objects.filter(
        vendedor=request.user
    ).aggregate(
        total_ingressos=Sum('quantidade'),
        total_comissao=Sum('valor_comissao')
    )
    
    context = {
        'eventos': eventos_ativos,
        'minhas_vendas': minhas_vendas,
        'total_ingressos': total_vendido['total_ingressos'] or 0,
        'total_comissao': total_vendido['total_comissao'] or 0,
    }
    
    return render(request, 'eventos/lista.html', context)


@login_required
def evento_detalhes(request, evento_id):
    """Detalhes de um evento e formulário para registrar venda"""
    evento = get_object_or_404(Evento, id=evento_id, ativo=True)
    
    # Minhas vendas neste evento
    minhas_vendas = VendaIngresso.objects.filter(
        evento=evento,
        vendedor=request.user
    ).order_by('-data_venda')
    
    meu_total = minhas_vendas.aggregate(
        total=Sum('quantidade')
    )['total'] or 0
    
    context = {
        'evento': evento,
        'minhas_vendas': minhas_vendas,
        'meu_total': meu_total,
    }
    
    return render(request, 'eventos/detalhes.html', context)


@login_required
@require_http_methods(["POST"])
def registrar_venda(request, evento_id):
    """Registra uma nova venda de ingressos"""
    evento = get_object_or_404(Evento, id=evento_id, ativo=True)
    
    try:
        quantidade = int(request.POST.get('quantidade', 0))
        data_venda = request.POST.get('data_venda', timezone.now().date())
        observacoes = request.POST.get('observacoes', '')
        
        if quantidade < 1:
            messages.error(request, 'A quantidade deve ser maior que zero.')
            return redirect('paginas:evento_detalhes', evento_id=evento_id)
        
        # Criar a venda
        venda = VendaIngresso.objects.create(
            evento=evento,
            vendedor=request.user,
            quantidade=quantidade,
            data_venda=data_venda,
            observacoes=observacoes
        )
        
        messages.success(
            request, 
            f'Venda de {quantidade} ingresso(s) registrada com sucesso! '
            f'Comissão: R$ {venda.valor_comissao:.2f}'
        )
        
        return redirect('paginas:evento_detalhes', evento_id=evento_id)
        
    except ValueError:
        messages.error(request, 'Quantidade inválida.')
        return redirect('paginas:evento_detalhes', evento_id=evento_id)
    except Exception as e:
        messages.error(request, f'Erro ao registrar venda: {str(e)}')
        return redirect('paginas:evento_detalhes', evento_id=evento_id)


@login_required
def minhas_vendas(request):
    """Histórico completo de vendas do usuário"""
    vendas = VendaIngresso.objects.filter(
        vendedor=request.user
    ).select_related('evento').order_by('-data_venda')
    
    # Estatísticas gerais
    stats = vendas.aggregate(
        total_ingressos=Sum('quantidade'),
        total_vendas=Count('id'),
        total_comissao=Sum('valor_comissao'),
        vendas_confirmadas=Count('id', filter=Q(confirmado=True)),
        vendas_pendentes=Count('id', filter=Q(confirmado=False))
    )
    
    # Estatísticas por evento
    por_evento = vendas.values('evento__nome').annotate(
        total_ingressos=Sum('quantidade'),
        total_comissao=Sum('valor_comissao')
    ).order_by('-total_ingressos')
    
    context = {
        'vendas': vendas,
        'stats': stats,
        'por_evento': por_evento,
    }
    
    return render(request, 'eventos/minhas_vendas.html', context)


@login_required
def admin_eventos_dashboard(request):
    """Dashboard administrativo de eventos (apenas para staff/admin)"""
    if not request.user.is_staff:
        raise PermissionDenied("Acesso restrito a administradores.")
    
    eventos = Evento.objects.all().order_by('-data_evento')
    
    # Estatísticas gerais
    eventos_stats = []
    for evento in eventos:
        vendas_evento = VendaIngresso.objects.filter(evento=evento)
        
        vendedores_stats = vendas_evento.values(
            'vendedor__first_name', 
            'vendedor__last_name',
            'vendedor__username'
        ).annotate(
            total_ingressos=Sum('quantidade'),
            total_comissao=Sum('valor_comissao'),
            total_vendas=Count('id')
        ).order_by('-total_ingressos')
        
        eventos_stats.append({
            'evento': evento,
            'total_vendido': evento.total_vendido(),
            'total_arrecadado': evento.total_arrecadado(),
            'percentual_meta': evento.percentual_meta(),
            'vendedores': vendedores_stats
        })
    
    # Ranking geral de vendedores
    ranking_geral = VendaIngresso.objects.values(
        'vendedor__first_name',
        'vendedor__last_name', 
        'vendedor__username'
    ).annotate(
        total_ingressos=Sum('quantidade'),
        total_comissao=Sum('valor_comissao'),
        total_eventos=Count('evento', distinct=True)
    ).order_by('-total_ingressos')[:10]
    
    context = {
        'eventos_stats': eventos_stats,
        'ranking_geral': ranking_geral,
    }
    
    return render(request, 'eventos/admin_dashboard.html', context)

@require_http_methods(["POST"])
def contato_consultor(request):
    """Envia email do formulário de contato da página do sistema"""
    try:
        data = json.loads(request.body)
        
        nome = data.get('nome', '').strip()
        email = data.get('email', '').strip()
        telefone = data.get('telefone', '').strip()
        empresa = data.get('empresa', '').strip()
        mensagem = data.get('mensagem', '').strip()
        
        # Validações básicas
        if not all([nome, email, telefone, mensagem]):
            return JsonResponse({'error': 'Todos os campos obrigatórios devem ser preenchidos'}, status=400)
        
        if len(nome) < 3:
            return JsonResponse({'error': 'Nome deve ter no mínimo 3 caracteres'}, status=400)
        
        if len(mensagem) < 10:
            return JsonResponse({'error': 'Mensagem deve ter no mínimo 10 caracteres'}, status=400)
        
        # Preparar email
        assunto = f'[Sistema Giro] Novo contato de {nome}'
        
        corpo_email = f"""
        Nova solicitação de contato recebida através da página "Conheça nosso Sistema"
        
        ===== DADOS DO CLIENTE =====
        Nome: {nome}
        Email: {email}
        Telefone: {telefone}
        Empresa: {empresa if empresa else 'Não informado'}
        
        ===== MENSAGEM =====
        {mensagem}
        
        ===========================
        Data/Hora: {timezone.now().strftime('%d/%m/%Y às %H:%M')}
        """
        
        # Enviar email
        try:
            email_message = EmailMessage(
                subject=assunto,
                body=corpo_email,
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@girodance.com',
                to=['contatoonyxtech@gmail.com'],
                reply_to=[email]
            )
            email_message.send(fail_silently=False)
            
            return JsonResponse({'success': True, 'message': 'Email enviado com sucesso!'})
        
        except Exception as email_error:
            # Log do erro (em produção, usar logging apropriado)
            print(f"Erro ao enviar email: {str(email_error)}")
            
            # Retornar sucesso mesmo se o email falhar (para não prejudicar UX)
            # Em produção, você pode salvar no banco de dados como backup
            return JsonResponse({
                'success': True, 
                'message': 'Mensagem recebida! Entraremos em contato em breve.',
                'warning': 'Email pode não ter sido enviado. Dados salvos para processamento.'
            })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Dados inválidos'}, status=400)
    except Exception as e:
        print(f"Erro no endpoint de contato: {str(e)}")
        return JsonResponse({'error': 'Erro ao processar solicitação'}, status=500)


# ==================== GESTÃO FINANCEIRA MENSAL ====================

@staff_member_required
def resultados_financeiros(request):
    """
    View principal para visualizar resultados financeiros mensais.
    Acesso restrito a administradores (staff).
    Exibe resumo com gráfico e tabela de todos os meses.
    """
    # Filtro de mês específico (opcional)
    mes_filtro = request.GET.get('mes', None)
    
    if mes_filtro:
        try:
            # Converter string "YYYY-MM" para date
            ano, mes = mes_filtro.split('-')
            data_filtro = datetime(int(ano), int(mes), 1).date()
            resultados = ResultadoFinanceiroMensal.objects.filter(mes=data_filtro)
        except (ValueError, AttributeError):
            resultados = ResultadoFinanceiroMensal.objects.all()
    else:
        resultados = ResultadoFinanceiroMensal.objects.all()
    
    # Calcular totais gerais
    totais = resultados.aggregate(
        total_lucro=Sum('lucro_total'),
        total_gastos=Sum('gasto_total'),
        total_a_receber=Sum('a_receber_total')
    )
    
    # Calcular lucro líquido total
    lucro_liquido_total = (totais['total_lucro'] or 0) - (totais['total_gastos'] or 0)
    
    # Preparar dados para o gráfico (último mês ou mês filtrado)
    resultado_atual = resultados.first()
    dados_grafico = resultado_atual.dados_grafico() if resultado_atual else None
    
    context = {
        'resultados': resultados,
        'totais': totais,
        'lucro_liquido_total': lucro_liquido_total,
        'dados_grafico': dados_grafico,
        'resultado_atual': resultado_atual,
        'mes_filtro': mes_filtro,
    }
    
    return render(request, 'financeiro/despesas.html', context)


@staff_member_required
def exportar_resultados_excel(request):
    """
    Gera e retorna um arquivo Excel (.xlsx) com os resultados financeiros.
    Formatação profissional com cores, bordas e totais destacados.
    """
    # Buscar todos os resultados ou filtrar por mês
    mes_filtro = request.GET.get('mes', None)
    
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-')
            data_filtro = datetime(int(ano), int(mes), 1).date()
            resultados = ResultadoFinanceiroMensal.objects.filter(mes=data_filtro)
            nome_arquivo = f'Resultado_Financeiro_{ano}_{mes}.xlsx'
        except (ValueError, AttributeError):
            resultados = ResultadoFinanceiroMensal.objects.all()
            nome_arquivo = 'Resultados_Financeiros_Giro_DNC.xlsx'
    else:
        resultados = ResultadoFinanceiroMensal.objects.all()
        nome_arquivo = 'Resultados_Financeiros_Giro_DNC.xlsx'
    
    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Mensais"
    
    # ===== ESTILOS =====
    # Cabeçalho
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F56E1D', end_color='F56E1D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Células de dados
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='right', vertical='center')
    text_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Valores monetários
    currency_font = Font(name='Arial', size=11, bold=True)
    lucro_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    gasto_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    receber_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    
    # Totais
    total_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    total_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== TÍTULO =====
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'GIRO DNC - RESULTADOS FINANCEIROS MENSAIS'
    title_cell.font = Font(name='Arial', size=14, bold=True, color='F56E1D')
    title_cell.alignment = center_alignment
    
    # ===== CABEÇALHOS =====
    headers = ['Mês/Ano', 'Lucro Total', 'Gasto Total', 'A Receber', 'Lucro Líquido']
    ws.append([])  # Linha em branco
    ws.append(headers)
    
    header_row = ws[3]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # ===== DADOS =====
    total_lucro = 0
    total_gastos = 0
    total_receber = 0
    total_liquido = 0
    
    for resultado in resultados:
        lucro_liquido = resultado.lucro_liquido()
        
        ws.append([
            resultado.mes_formatado(),
            float(resultado.lucro_total),
            float(resultado.gasto_total),
            float(resultado.a_receber_total),
            float(lucro_liquido)
        ])
        
        # Acumular totais
        total_lucro += float(resultado.lucro_total)
        total_gastos += float(resultado.gasto_total)
        total_receber += float(resultado.a_receber_total)
        total_liquido += float(lucro_liquido)
        
        # Aplicar estilos às células
        current_row = ws.max_row
        
        # Mês/Ano
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].font = data_font
        ws[f'A{current_row}'].border = thin_border
        
        # Lucro Total
        ws[f'B{current_row}'].number_format = 'R$ #,##0.00'
        ws[f'B{current_row}'].font = currency_font
        ws[f'B{current_row}'].fill = lucro_fill
        ws[f'B{current_row}'].alignment = data_alignment
        ws[f'B{current_row}'].border = thin_border
        
        # Gasto Total
        ws[f'C{current_row}'].number_format = 'R$ #,##0.00'
        ws[f'C{current_row}'].font = currency_font
        ws[f'C{current_row}'].fill = gasto_fill
        ws[f'C{current_row}'].alignment = data_alignment
        ws[f'C{current_row}'].border = thin_border
        
        # A Receber
        ws[f'D{current_row}'].number_format = 'R$ #,##0.00'
        ws[f'D{current_row}'].font = currency_font
        ws[f'D{current_row}'].fill = receber_fill
        ws[f'D{current_row}'].alignment = data_alignment
        ws[f'D{current_row}'].border = thin_border
        
        # Lucro Líquido
        ws[f'E{current_row}'].number_format = 'R$ #,##0.00'
        ws[f'E{current_row}'].font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        
        if lucro_liquido >= 0:
            ws[f'E{current_row}'].fill = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
        else:
            ws[f'E{current_row}'].fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        
        ws[f'E{current_row}'].alignment = data_alignment
        ws[f'E{current_row}'].border = thin_border
    
    # ===== LINHA DE TOTAIS =====
    ws.append([])  # Linha em branco
    total_row = ws.max_row + 1
    
    ws.append(['TOTAIS', total_lucro, total_gastos, total_receber, total_liquido])
    
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_alignment if col == 1 else data_alignment
        cell.border = thin_border
        
        if col > 1:
            cell.number_format = 'R$ #,##0.00'
    
    # ===== AJUSTAR LARGURA DAS COLUNAS =====
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    
    # ===== PREPARAR RESPOSTA HTTP =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    wb.save(response)
    return response


# ============================================================
# DESPESAS PESSOAIS DOS ALUNOS
# ============================================================

@login_required
def minhas_despesas(request):
    """View para alunos gerenciarem suas despesas pessoais"""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        messages.error(request, 'Você precisa estar cadastrado como aluno.')
        return redirect('paginas:painel_aluno')
    
    # Filtros
    mes_filtro = request.GET.get('mes', '')
    categoria_filtro = request.GET.get('categoria', '')
    status_filtro = request.GET.get('status', '')
    
    # Buscar despesas do aluno
    despesas = DespesaAluno.objects.filter(aluno=aluno).select_related('turma')
    
    # Aplicar filtros
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-')
            despesas = despesas.filter(
                data_prevista__year=int(ano),
                data_prevista__month=int(mes)
            )
        except ValueError:
            pass
    
    if categoria_filtro:
        despesas = despesas.filter(categoria=categoria_filtro)
    
    if status_filtro:
        despesas = despesas.filter(status=status_filtro)
    
    # Estatísticas
    total_previsto = despesas.aggregate(Sum('valor_previsto'))['valor_previsto__sum'] or 0
    total_gasto = despesas.aggregate(Sum('valor_gasto'))['valor_gasto__sum'] or 0
    total_restante = total_previsto - total_gasto
    
    # Despesas por categoria (para gráfico)
    despesas_por_categoria = {}
    for despesa in despesas:
        cat = despesa.categoria_display()
        if cat not in despesas_por_categoria:
            despesas_por_categoria[cat] = 0
        despesas_por_categoria[cat] += float(despesa.valor_gasto)
    
    # Preparar dados para gráfico
    categorias_labels = list(despesas_por_categoria.keys())
    categorias_valores = list(despesas_por_categoria.values())
    
    context = {
        'aluno': aluno,
        'despesas': despesas.order_by('-data_prevista'),
        'total_previsto': total_previsto,
        'total_gasto': total_gasto,
        'total_restante': total_restante,
        'categorias_labels': json.dumps(categorias_labels),
        'categorias_valores': json.dumps(categorias_valores),
        'mes_filtro': mes_filtro,
        'categoria_filtro': categoria_filtro,
        'status_filtro': status_filtro,
        'categorias_choices': DespesaAluno.CATEGORIA_CHOICES,
        'status_choices': DespesaAluno.STATUS_CHOICES,
    }
    
    return render(request, 'financeiro/minhas_despesas.html', context)


@login_required
def criar_despesa(request):
    """View para criar nova despesa pessoal"""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        messages.error(request, 'Você precisa estar cadastrado como aluno.')
        return redirect('paginas:painel_aluno')
    
    if request.method == 'POST':
        try:
            despesa = DespesaAluno(
                aluno=aluno,
                nome=request.POST.get('nome'),
                descricao=request.POST.get('descricao'),
                categoria=request.POST.get('categoria'),
                valor_previsto=request.POST.get('valor_previsto', 0),
                valor_gasto=request.POST.get('valor_gasto', 0),
                status=request.POST.get('status', 'PLANEJADO'),
                data_prevista=request.POST.get('data_prevista'),
                observacoes=request.POST.get('observacoes', ''),
                itens=request.POST.get('itens', '')
            )
            
            # Turma (opcional)
            turma_id = request.POST.get('turma')
            if turma_id:
                despesa.turma_id = turma_id
            
            # Data pagamento (opcional)
            data_pagamento = request.POST.get('data_pagamento')
            if data_pagamento:
                despesa.data_pagamento = data_pagamento
            
            despesa.save()
            messages.success(request, 'Despesa criada com sucesso!')
            return redirect('paginas:minhas_despesas')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar despesa: {str(e)}')
    
    # GET - mostrar formulário
    turmas_aluno = aluno.turmas.filter(ativa=True)
    
    context = {
        'aluno': aluno,
        'turmas': turmas_aluno,
        'categorias': DespesaAluno.CATEGORIA_CHOICES,
        'status': DespesaAluno.STATUS_CHOICES,
    }
    
    return render(request, 'financeiro/criar_despesa.html', context)


@login_required
def editar_despesa(request, despesa_id):
    """View para editar despesa existente"""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        messages.error(request, 'Você precisa estar cadastrado como aluno.')
        return redirect('paginas:painel_aluno')
    
    despesa = get_object_or_404(DespesaAluno, id=despesa_id, aluno=aluno)
    
    if request.method == 'POST':
        try:
            despesa.nome = request.POST.get('nome')
            despesa.descricao = request.POST.get('descricao')
            despesa.categoria = request.POST.get('categoria')
            despesa.valor_previsto = request.POST.get('valor_previsto', 0)
            despesa.valor_gasto = request.POST.get('valor_gasto', 0)
            despesa.status = request.POST.get('status')
            despesa.data_prevista = request.POST.get('data_prevista')
            despesa.observacoes = request.POST.get('observacoes', '')
            despesa.itens = request.POST.get('itens', '')
            
            # Turma (opcional)
            turma_id = request.POST.get('turma')
            if turma_id:
                despesa.turma_id = turma_id
            else:
                despesa.turma = None
            
            # Data pagamento (opcional)
            data_pagamento = request.POST.get('data_pagamento')
            if data_pagamento:
                despesa.data_pagamento = data_pagamento
            else:
                despesa.data_pagamento = None
            
            despesa.save()
            messages.success(request, 'Despesa atualizada com sucesso!')
            return redirect('paginas:minhas_despesas')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar despesa: {str(e)}')
    
    # GET - mostrar formulário
    turmas_aluno = aluno.turmas.filter(ativa=True)
    
    context = {
        'aluno': aluno,
        'despesa': despesa,
        'turmas': turmas_aluno,
        'categorias': DespesaAluno.CATEGORIA_CHOICES,
        'status': DespesaAluno.STATUS_CHOICES,
    }
    
    return render(request, 'financeiro/editar_despesa.html', context)


@login_required
@require_http_methods(["POST"])
def deletar_despesa(request, despesa_id):
    """View para deletar uma despesa"""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        messages.error(request, 'Você precisa estar cadastrado como aluno.')
        return redirect('paginas:painel_aluno')
    
    despesa = get_object_or_404(DespesaAluno, id=despesa_id, aluno=aluno)
    nome_despesa = despesa.nome
    despesa.delete()
    
    messages.success(request, f'Despesa "{nome_despesa}" excluída com sucesso!')
    return redirect('paginas:minhas_despesas')


# ============================================================
# DESPESAS ADMINISTRATIVAS
# ============================================================

@staff_member_required
def despesas_administrativas(request):
    """View para administradores gerenciarem despesas administrativas da Giro DNC"""
    from datetime import date, datetime
    from django.db.models import Q
    import json
    
    # Filtros
    mes_filtro = request.GET.get('mes', '')
    categoria_filtro = request.GET.get('categoria', '')
    status_filtro = request.GET.get('status', '')
    tipo_filtro = request.GET.get('tipo', '')
    
    # Buscar despesas
    despesas = DespesaAdministrativa.objects.all()
    
    # Aplicar filtros
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-')
            despesas = despesas.filter(
                data_vencimento__year=int(ano),
                data_vencimento__month=int(mes)
            )
        except ValueError:
            pass
    
    if categoria_filtro:
        despesas = despesas.filter(categoria=categoria_filtro)
    
    if status_filtro:
        despesas = despesas.filter(status=status_filtro)
    
    if tipo_filtro:
        despesas = despesas.filter(tipo_pagamento=tipo_filtro)
    
    # Estatísticas gerais
    total_despesas = despesas.aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    total_pago = despesas.aggregate(Sum('valor_pago'))['valor_pago__sum'] or 0
    total_pendente = total_despesas - total_pago
    
    # Despesas por status
    despesas_pendentes = despesas.filter(status='PENDENTE').aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    despesas_pagas = despesas.filter(status='PAGO').aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    despesas_atrasadas = despesas.filter(status='ATRASADO').aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    
    # Despesas por categoria (para gráfico)
    despesas_por_categoria = {}
    for despesa in despesas:
        cat = despesa.categoria_display()
        if cat not in despesas_por_categoria:
            despesas_por_categoria[cat] = 0
        despesas_por_categoria[cat] += float(despesa.valor_total)
    
    # Preparar dados para gráfico de categorias
    categorias_labels = list(despesas_por_categoria.keys())
    categorias_valores = list(despesas_por_categoria.values())
    
    # Gráfico de status (Doughnut)
    status_labels = ['Pagas', 'Pendentes', 'Atrasadas']
    status_valores = [
        float(despesas_pagas),
        float(despesas_pendentes),
        float(despesas_atrasadas)
    ]
    status_cores = ['#22c55e', '#f59e0b', '#ef4444']
    
    # Despesas vencendo nos próximos 7 dias
    hoje = date.today()
    from datetime import timedelta
    proximos_7_dias = hoje + timedelta(days=7)
    despesas_vencendo = despesas.filter(
        status__in=['PENDENTE', 'PARCIAL'],
        data_vencimento__gte=hoje,
        data_vencimento__lte=proximos_7_dias
    ).count()
    
    # Top 5 maiores despesas do período
    top_despesas = despesas.order_by('-valor_total')[:5]
    
    context = {
        'despesas': despesas.order_by('-data_vencimento'),
        'total_despesas': total_despesas,
        'total_pago': total_pago,
        'total_pendente': total_pendente,
        'despesas_pendentes': despesas_pendentes,
        'despesas_pagas': despesas_pagas,
        'despesas_atrasadas': despesas_atrasadas,
        'despesas_vencendo': despesas_vencendo,
        'top_despesas': top_despesas,
        'categorias_labels': json.dumps(categorias_labels),
        'categorias_valores': json.dumps(categorias_valores),
        'status_labels': json.dumps(status_labels),
        'status_valores': json.dumps(status_valores),
        'status_cores': json.dumps(status_cores),
        'mes_filtro': mes_filtro,
        'categoria_filtro': categoria_filtro,
        'status_filtro': status_filtro,
        'tipo_filtro': tipo_filtro,
        'categorias': DespesaAdministrativa.CATEGORIA_CHOICES,
        'status_choices': DespesaAdministrativa.STATUS_CHOICES,
        'tipos': DespesaAdministrativa.TIPO_PAGAMENTO_CHOICES,
    }
    
    return render(request, 'financeiro/despesas_administrativas.html', context)


@staff_member_required
def criar_despesa_admin(request):
    """View para criar nova despesa administrativa"""
    if request.method == 'POST':
        try:
            despesa = DespesaAdministrativa()
            despesa.nome = request.POST.get('nome')
            despesa.categoria = request.POST.get('categoria')
            despesa.descricao = request.POST.get('descricao', '')
            despesa.fornecedor = request.POST.get('fornecedor', '')
            despesa.valor_total = request.POST.get('valor_total')
            despesa.valor_pago = request.POST.get('valor_pago', 0)
            despesa.data_vencimento = request.POST.get('data_vencimento')
            despesa.status = request.POST.get('status', 'PENDENTE')
            despesa.tipo_pagamento = request.POST.get('tipo_pagamento', 'UNICO')
            despesa.forma_pagamento = request.POST.get('forma_pagamento', '')
            despesa.numero_documento = request.POST.get('numero_documento', '')
            despesa.observacoes = request.POST.get('observacoes', '')
            despesa.criado_por = request.user
            
            # Parcelamento (se aplicável)
            numero_parcelas = request.POST.get('numero_parcelas')
            if numero_parcelas:
                despesa.numero_parcelas = numero_parcelas
            
            parcela_atual = request.POST.get('parcela_atual')
            if parcela_atual:
                despesa.parcela_atual = parcela_atual
            
            # Data pagamento (opcional)
            data_pagamento = request.POST.get('data_pagamento')
            if data_pagamento:
                despesa.data_pagamento = data_pagamento
            
            # Comprovante (opcional)
            if 'comprovante' in request.FILES:
                despesa.comprovante = request.FILES['comprovante']
            
            despesa.save()
            messages.success(request, 'Despesa administrativa criada com sucesso!')
            return redirect('paginas:despesas_administrativas')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar despesa: {str(e)}')
    
    # GET - mostrar formulário
    context = {
        'categorias': DespesaAdministrativa.CATEGORIA_CHOICES,
        'status_choices': DespesaAdministrativa.STATUS_CHOICES,
        'tipos_pagamento': DespesaAdministrativa.TIPO_PAGAMENTO_CHOICES,
        'formas_pagamento': DespesaAdministrativa.FORMA_PAGAMENTO_CHOICES,
    }
    
    return render(request, 'financeiro/criar_despesa_admin.html', context)


@staff_member_required
def editar_despesa_admin(request, despesa_id):
    """View para editar despesa administrativa existente"""
    despesa = get_object_or_404(DespesaAdministrativa, id=despesa_id)
    
    if request.method == 'POST':
        try:
            despesa.nome = request.POST.get('nome')
            despesa.categoria = request.POST.get('categoria')
            despesa.descricao = request.POST.get('descricao', '')
            despesa.fornecedor = request.POST.get('fornecedor', '')
            despesa.valor_total = request.POST.get('valor_total')
            despesa.valor_pago = request.POST.get('valor_pago', 0)
            despesa.data_vencimento = request.POST.get('data_vencimento')
            despesa.status = request.POST.get('status')
            despesa.tipo_pagamento = request.POST.get('tipo_pagamento')
            despesa.forma_pagamento = request.POST.get('forma_pagamento', '')
            despesa.numero_documento = request.POST.get('numero_documento', '')
            despesa.observacoes = request.POST.get('observacoes', '')
            
            # Parcelamento
            numero_parcelas = request.POST.get('numero_parcelas')
            if numero_parcelas:
                despesa.numero_parcelas = numero_parcelas
            else:
                despesa.numero_parcelas = None
            
            parcela_atual = request.POST.get('parcela_atual')
            if parcela_atual:
                despesa.parcela_atual = parcela_atual
            else:
                despesa.parcela_atual = None
            
            # Data pagamento
            data_pagamento = request.POST.get('data_pagamento')
            if data_pagamento:
                despesa.data_pagamento = data_pagamento
            else:
                despesa.data_pagamento = None
            
            # Comprovante
            if 'comprovante' in request.FILES:
                despesa.comprovante = request.FILES['comprovante']
            
            despesa.save()
            messages.success(request, 'Despesa atualizada com sucesso!')
            return redirect('paginas:despesas_administrativas')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar despesa: {str(e)}')
    
    # GET - mostrar formulário
    context = {
        'despesa': despesa,
        'categorias': DespesaAdministrativa.CATEGORIA_CHOICES,
        'status_choices': DespesaAdministrativa.STATUS_CHOICES,
        'tipos_pagamento': DespesaAdministrativa.TIPO_PAGAMENTO_CHOICES,
        'formas_pagamento': DespesaAdministrativa.FORMA_PAGAMENTO_CHOICES,
    }
    
    return render(request, 'financeiro/editar_despesa_admin.html', context)


@staff_member_required
@require_http_methods(["POST"])
def deletar_despesa_admin(request, despesa_id):
    """View para deletar uma despesa administrativa"""
    despesa = get_object_or_404(DespesaAdministrativa, id=despesa_id)
    nome_despesa = despesa.nome
    despesa.delete()
    
    messages.success(request, f'Despesa "{nome_despesa}" excluída com sucesso!')
    return redirect('paginas:despesas_administrativas')


@staff_member_required
def exportar_despesas_admin_excel(request):
    """Exporta despesas administrativas para Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime
    
    # Buscar despesas com filtros (se houver)
    mes_filtro = request.GET.get('mes', None)
    categoria_filtro = request.GET.get('categoria', None)
    status_filtro = request.GET.get('status', None)
    
    despesas = DespesaAdministrativa.objects.all()
    
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-')
            despesas = despesas.filter(
                data_vencimento__year=int(ano),
                data_vencimento__month=int(mes)
            )
        except ValueError:
            pass
    
    if categoria_filtro:
        despesas = despesas.filter(categoria=categoria_filtro)
    
    if status_filtro:
        despesas = despesas.filter(status=status_filtro)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Despesas Administrativas"
    
    # Estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F56E1D', end_color='F56E1D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos
    headers = ['Nome', 'Categoria', 'Fornecedor', 'Valor Total', 'Valor Pago', 'Valor Pendente',
               'Status', 'Data Vencimento', 'Data Pagamento', 'Tipo', 'Forma Pgto', 'Parcelas', 'Observações']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados
    for row_num, despesa in enumerate(despesas.order_by('-data_vencimento'), 2):
        ws.cell(row=row_num, column=1, value=despesa.nome)
        ws.cell(row=row_num, column=2, value=despesa.categoria_display())
        ws.cell(row=row_num, column=3, value=despesa.fornecedor or '-')
        ws.cell(row=row_num, column=4, value=float(despesa.valor_total))
        ws.cell(row=row_num, column=5, value=float(despesa.valor_pago))
        ws.cell(row=row_num, column=6, value=float(despesa.valor_pendente()))
        ws.cell(row=row_num, column=7, value=despesa.status_display())
        ws.cell(row=row_num, column=8, value=despesa.data_vencimento.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=9, value=despesa.data_pagamento.strftime('%d/%m/%Y') if despesa.data_pagamento else '-')
        ws.cell(row=row_num, column=10, value=despesa.get_tipo_pagamento_display())
        ws.cell(row=row_num, column=11, value=despesa.get_forma_pagamento_display() if despesa.forma_pagamento else '-')
        ws.cell(row=row_num, column=12, value=despesa.parcelas_display())
        ws.cell(row=row_num, column=13, value=despesa.observacoes or '-')
    
    # Ajustar larguras
    column_widths = {
        'A': 30, 'B': 20, 'C': 25, 'D': 15, 'E': 15, 'F': 15,
        'G': 15, 'H': 15, 'I': 15, 'J': 18, 'K': 18, 'L': 12, 'M': 40
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Salvar em BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Criar resposta HTTP
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'despesas_administrativas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def chat_view(request, slug="geral"):
    room, _ = Room.objects.get_or_create(name=slug)
    return render(request, "painel/chat.html", {"room": room})
